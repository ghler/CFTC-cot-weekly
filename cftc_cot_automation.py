#!/usr/bin/env python3
"""
CFTC COT Data Automation
- Downloads latest weekly COT data (Futures Only, XLS format)
- Extracts specified market codes (16 commodities)
- Keeps ALL columns (A through DV, 126 columns) exactly as in source
- Retains only the most recent 3 weeks per market code
- Runs weekly via GitHub Actions (Friday after 3:30 PM ET when CFTC publishes)
"""

import os
import sys
import zipfile
import tempfile
import shutil
from datetime import datetime
from pathlib import Path
import requests
import pandas as pd
from openpyxl import load_workbook, Workbook

# ========== CONFIGURATION ==========
MASTER_FILE = Path("data/cot_master.xlsx")      # Master file (committed to repo)
SHEET_NAME = "COT_Multi_Commodity"              # Sheet name in master

# 16 Market Codes to track (as strings to preserve leading zeros)
TARGET_CODES = [
    "043602", "084691", "088691", "090741", "092741",
    "095741", "096742", "097741", "098662", "099741",
    "112741", "134741", "232741", "299741", "067411", "067651"
]

# Number of recent weeks to keep PER market code
WEEKS_TO_KEEP = 3

# CFTC URL pattern (yearly zip files)
BASE_URL = "https://www.cftc.gov/files/dea/history/dea_fut_xls_{year}.zip"
# ====================================


def get_latest_cftc_url() -> str:
    """Determine the latest available CFTC zip URL."""
    now = datetime.now()
    for year in [now.year, now.year - 1]:
        url = BASE_URL.format(year=year)
        try:
            resp = requests.head(url, timeout=10)
            if resp.status_code == 200:
                return url
        except Exception:
            pass
    raise RuntimeError("Could not find available CFTC data file")


def download_and_extract(url: str) -> Path:
    """Download zip and extract annual.xls to temp dir."""
    tmpdir = Path(tempfile.mkdtemp(prefix="cftc_"))
    zip_path = tmpdir / "data.zip"
    
    print(f"Downloading {url}...")
    resp = requests.get(url, timeout=60, stream=True)
    resp.raise_for_status()
    zip_path.write_bytes(resp.content)
    
    print("Extracting...")
    with zipfile.ZipFile(zip_path, 'r') as zf:
        xls_name = "annual.xls"
        if xls_name not in zf.namelist():
            raise FileNotFoundError(f"{xls_name} not in zip: {zf.namelist()}")
        zf.extract(xls_name, tmpdir)
    
    return tmpdir / xls_name


def read_cftc_xls(xls_path: Path) -> pd.DataFrame:
    """Read CFTC .xls file (requires xlrd). Returns ALL columns."""
    print(f"Reading {xls_path}...")
    df = pd.read_excel(xls_path, engine="xlrd")
    print(f"  Loaded {len(df)} rows, {len(df.columns)} columns")
    return df


def filter_target_codes(df: pd.DataFrame) -> pd.DataFrame:
    """Extract rows for all target market codes, keeping ALL columns."""
    # Match by CFTC_Contract_Market_Code (exact string match)
    mask = df["CFTC_Contract_Market_Code"].astype(str).str.strip().isin(TARGET_CODES)
    result = df.loc[mask].copy()
    
    # Parse date column for sorting
    result["Report_Date_as_MM_DD_YYYY"] = pd.to_datetime(
        result["Report_Date_as_MM_DD_YYYY"], errors="coerce"
    )
    
    # Sort by market code, then by date (newest first)
    result = result.sort_values(
        ["CFTC_Contract_Market_Code", "Report_Date_as_MM_DD_YYYY"],
        ascending=[True, False]
    )
    
    print(f"  Found {len(result)} total rows for {len(TARGET_CODES)} market codes")
    for code in TARGET_CODES:
        count = len(result[result["CFTC_Contract_Market_Code"].astype(str).str.strip() == code])
        print(f"    {code}: {count} weeks")
    
    return result


def keep_recent_weeks_per_code(df: pd.DataFrame, weeks: int = WEEKS_TO_KEEP) -> pd.DataFrame:
    """Keep only the most recent N weeks per market code."""
    # Group by market code and take top N rows (already sorted newest first)
    kept = df.groupby("CFTC_Contract_Market_Code", group_keys=False).head(weeks)
    
    # Sort final output: by market code, then by date ascending (oldest first within each code)
    kept = kept.sort_values(
        ["CFTC_Contract_Market_Code", "Report_Date_as_MM_DD_YYYY"],
        ascending=[True, True]
    ).reset_index(drop=True)
    
    print(f"  Keeping most recent {weeks} weeks per code → {len(kept)} total rows")
    return kept


def load_master_file() -> pd.DataFrame:
    """Load existing master file or return empty DataFrame with correct columns."""
    if MASTER_FILE.exists():
        try:
            df = pd.read_excel(MASTER_FILE, sheet_name=SHEET_NAME, engine="openpyxl")
            # Ensure date column is datetime
            if "Report_Date_as_MM_DD_YYYY" in df.columns:
                df["Report_Date_as_MM_DD_YYYY"] = pd.to_datetime(
                    df["Report_Date_as_MM_DD_YYYY"], errors="coerce"
                )
            print(f"Loaded master: {len(df)} existing rows")
            return df
        except Exception as e:
            print(f"Warning: Could not read master file: {e}")
    # Return empty df with no predefined columns — we'll use source columns
    return pd.DataFrame()


def append_new_rows(master_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """
    Append new rows not already in master.
    Deduplication key: (Market Code, Report Date)
    """
    if master_df.empty:
        return new_df
    
    # Build set of existing (code, date) pairs
    master_df["_code"] = master_df["CFTC_Contract_Market_Code"].astype(str).str.strip()
    master_df["_date"] = master_df["Report_Date_as_MM_DD_YYYY"].dt.date
    existing_pairs = set(zip(master_df["_code"], master_df["_date"]))
    
    # Filter new_df for pairs not in master
    new_df["_code"] = new_df["CFTC_Contract_Market_Code"].astype(str).str.strip()
    new_df["_date"] = new_df["Report_Date_as_MM_DD_YYYY"].dt.date
    new_pairs = list(zip(new_df["_code"], new_df["_date"]))
    
    mask = [pair not in existing_pairs for pair in new_pairs]
    truly_new = new_df[mask].drop(columns=["_code", "_date"])
    
    # Clean up master temp columns
    master_df = master_df.drop(columns=["_code", "_date"])
    
    if truly_new.empty:
        print("No new rows to append (all code+date pairs already in master)")
        return master_df
    
    print(f"Appending {len(truly_new)} new rows:")
    for _, row in truly_new.iterrows():
        code = row["CFTC_Contract_Market_Code"]
        date = row["Report_Date_as_MM_DD_YYYY"].date()
        oi = row.get("Open_Interest_All", "N/A")
        print(f"  + {code} | {date} | OI={oi}")
    
    # Combine and sort: by market code, then date ascending
    combined = pd.concat([master_df, truly_new], ignore_index=True)
    combined = combined.sort_values(
        ["CFTC_Contract_Market_Code", "Report_Date_as_MM_DD_YYYY"],
        ascending=[True, True]
    ).reset_index(drop=True)
    
    return combined


def trim_master_to_recent_weeks(master_df: pd.DataFrame, weeks: int = WEEKS_TO_KEEP) -> pd.DataFrame:
    """
    After appending, trim master to only keep most recent N weeks per market code.
    This handles the case where we accumulate data over time.
    """
    if master_df.empty:
        return master_df
    
    # Sort by code, then date descending (newest first)
    master_df = master_df.sort_values(
        ["CFTC_Contract_Market_Code", "Report_Date_as_MM_DD_YYYY"],
        ascending=[True, False]
    )
    
    # Keep top N per code
    trimmed = master_df.groupby("CFTC_Contract_Market_Code", group_keys=False).head(weeks)
    
    # Final sort: by code, then date ascending (oldest first within each code)
    trimmed = trimmed.sort_values(
        ["CFTC_Contract_Market_Code", "Report_Date_as_MM_DD_YYYY"],
        ascending=[True, True]
    ).reset_index(drop=True)
    
    removed = len(master_df) - len(trimmed)
    if removed > 0:
        print(f"  Trimmed {removed} old rows (keeping most recent {weeks} weeks per code)")
    
    return trimmed


def save_master_file(df: pd.DataFrame):
    """Save master file with ALL columns preserved, header exactly as source."""
    MASTER_FILE.parent.mkdir(parents=True, exist_ok=True)
    
    # Get column order from the dataframe (should match source)
    all_columns = list(df.columns)
    
    # Write with openpyxl
    if MASTER_FILE.exists():
        wb = load_workbook(MASTER_FILE)
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    
    ws = wb.create_sheet(SHEET_NAME)
    
    # Header row (exact column names from source)
    for col_idx, col_name in enumerate(all_columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(all_columns, 1):
            val = row[col_name]
            # Handle NaN/None and datetime
            if pd.isna(val):
                val = None
            elif isinstance(val, (pd.Timestamp, datetime)):
                val = val.date() if isinstance(val, pd.Timestamp) else val
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    # Auto-filter on all columns
    last_col_letter = _col_num_to_letter(len(all_columns))
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(df)+1}"
    
    wb.save(MASTER_FILE)
    print(f"Saved master: {MASTER_FILE} ({len(df)} rows, {len(all_columns)} columns)")


def _col_num_to_letter(n: int) -> str:
    """Convert column number to Excel letter (1=A, 27=AA, etc.)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def cleanup_temp(tmpdir: Path):
    """Remove temp directory."""
    shutil.rmtree(tmpdir, ignore_errors=True)


def main():
    print(f"=== CFTC COT Automation - {datetime.now().isoformat()} ===")
    print(f"Target codes: {', '.join(TARGET_CODES)}")
    print(f"Retaining: most recent {WEEKS_TO_KEEP} weeks per code")
    print(f"Columns: ALL ({'will be determined from source'})")
    
    tmpdir = None
    try:
        # 1. Get latest data URL
        url = get_latest_cftc_url()
        print(f"Using: {url}")
        
        # 2. Download & extract
        xls_path = download_and_extract(url)
        tmpdir = xls_path.parent
        
        # 3. Read & filter
        df = read_cftc_xls(xls_path)
        new_data = filter_target_codes(df)
        
        if new_data.empty:
            print("ERROR: No data found for target codes!")
            sys.exit(1)
        
        # 4. Keep only recent weeks from new download
        new_data = keep_recent_weeks_per_code(new_data, WEEKS_TO_KEEP)
        
        # 5. Load master & append new rows
        master = load_master_file()
        updated = append_new_rows(master, new_data)
        
        # 6. Trim master to recent weeks per code (enforces retention policy)
        updated = trim_master_to_recent_weeks(updated, WEEKS_TO_KEEP)
        
        # 7. Save
        save_master_file(updated)
        
        print("=== DONE ===")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        if tmpdir:
            cleanup_temp(tmpdir)


if __name__ == "__main__":
    main()